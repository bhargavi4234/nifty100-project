import os
import sqlite3

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .engine import apply_filters

# ==========================================================
# Configuration
# ==========================================================

CONFIG_FILE = "config/screener_config.yaml"
OUTPUT_FILE = "output/screener_output.xlsx"

MAX_RESULTS = 50


def load_presets():
    """Load all screener preset thresholds from YAML."""

    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f) or {}

    return config.get("presets", {})


PRESETS = load_presets()


# ==========================================================
# Run Preset
# ==========================================================


def run_preset(df, preset_name):
    """
    Apply a predefined screener preset.

    Results are sorted by composite quality score and
    limited to the top 50 records.
    """

    if preset_name not in PRESETS:
        raise ValueError(f"Unknown preset: {preset_name}")

    result = apply_filters(df, PRESETS[preset_name])

    # Keep the best 50 results
    result = (
        result.sort_values(by="composite_quality_score", ascending=False)
        .head(MAX_RESULTS)
        .reset_index(drop=True)
    )

    return result


# ==========================================================
# Main
# ==========================================================

if __name__ == "__main__":

    conn = sqlite3.connect("data/db/nifty100.db")

    query = """
    SELECT

        fr.*,

        mc.market_cap_crore,
        mc.pe_ratio,
        mc.pb_ratio,
        mc.dividend_yield_pct,

        pl.sales,
        pl.net_profit,

        s.broad_sector

    FROM financial_ratios fr

    LEFT JOIN market_cap mc
        ON fr.company_id = mc.company_id
        AND SUBSTR(fr.year, -4) =
            CAST(mc.year AS TEXT)

    LEFT JOIN profitandloss pl
        ON fr.company_id = pl.company_id
        AND fr.year = pl.year

    LEFT JOIN sectors s
        ON fr.company_id = s.company_id
    """

    df = pd.read_sql(query, conn)

    conn.close()

    os.makedirs("output", exist_ok=True)

    writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")

    print("=" * 70)
    print("Financial Screener Presets")
    print("=" * 70)

    columns = [
        "company_id",
        "year",
        "broad_sector",
        "composite_quality_score",
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "free_cash_flow_cr",
        "cash_from_operations_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "debt_to_equity",
        "interest_coverage",
        "market_cap_crore",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield_pct",
        "sales",
    ]

    for preset_name in PRESETS:

        print("\n" + "=" * 70)
        print(f"Testing: {preset_name}")

        result = run_preset(df, preset_name)

        available = [c for c in columns if c in result.columns]

        result[available].to_excel(writer, sheet_name=preset_name[:31], index=False)

        print(f"Preset   : {preset_name}")

        print(f"Companies: {len(result)}")

        if len(result) > 0:
            print(result[["company_id", "year", "composite_quality_score"]].head(5))
        else:
            print("No companies matched.")

    writer.close()


# ==========================================================
# Colour-code workbook
# ==========================================================

wb = load_workbook(OUTPUT_FILE)

green = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")

red = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")


metric_columns = {
    "return_on_equity_pct": "roe_min",
    "debt_to_equity": "debt_equity_max",
    "free_cash_flow_cr": "fcf_min",
    "revenue_cagr_5yr": "revenue_cagr_5yr_min",
    "pat_cagr_5yr": "pat_cagr_5yr_min",
    "operating_profit_margin_pct": "opm_min",
    "market_cap_crore": "market_cap_min",
    "sales": "sales_min",
    "net_profit": "net_profit_min",
    "pe_ratio": "pe_max",
    "pb_ratio": "pb_max",
    "dividend_yield_pct": "dividend_yield_min",
}


for sheet in wb.sheetnames:

    ws = wb[sheet]

    preset = PRESETS[sheet]

    headers = [cell.value for cell in ws[1]]

    for col_name, filter_name in metric_columns.items():

        if (
            col_name not in headers
            or filter_name not in preset
            or preset[filter_name] is None
        ):
            continue

        col = headers.index(col_name) + 1

        threshold = preset[filter_name]

        for row in range(2, ws.max_row + 1):

            value = ws.cell(row, col).value

            if value is None:
                continue

            try:
                value = float(value)
            except (TypeError, ValueError):
                continue

            if filter_name.endswith("_max"):

                if value <= threshold:
                    ws.cell(row, col).fill = green
                else:
                    ws.cell(row, col).fill = red

            else:

                if value >= threshold:
                    ws.cell(row, col).fill = green
                else:
                    ws.cell(row, col).fill = red


wb.save(OUTPUT_FILE)


print("\nExcel exported to:")
print(OUTPUT_FILE)
print("\n" + "=" * 70)
print("All presets tested successfully.")
print("=" * 70)

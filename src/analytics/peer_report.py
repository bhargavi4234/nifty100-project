import os
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

DB_PATH = "data/db/nifty100.db"
OUTPUT_FILE = "output/peer_comparison.xlsx"

os.makedirs("output", exist_ok=True)


# ==========================================================
# Excel Colours
# ==========================================================

GREEN = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

YELLOW = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
    end_color="FFEB9C",
)

RED = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)

GOLD = PatternFill(
    fill_type="solid",
    start_color="FFD966",
    end_color="FFD966",
)


# ==========================================================
# 20 Metric Columns
# ==========================================================

METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "cash_from_operations_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "composite_quality_score",
    "market_cap_crore",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "sales",
    "net_profit",
    "dividend_payout",
]


# ==========================================================
# Metrics where LOWER is better
# ==========================================================

LOWER_IS_BETTER = {
    "debt_to_equity",
    "pe_ratio",
    "pb_ratio",
}


# ==========================================================
# Calculate percentile rank
# ==========================================================


def calculate_percentile(series, inverse=False):
    "Calculate percentile."

    values = pd.to_numeric(
        series,
        errors="coerce",
    )

    percentile = values.rank(
        pct=True,
        method="average",
    )

    if inverse:
        percentile = 1 - percentile

    return percentile.round(4)


# ==========================================================
# Main
# ==========================================================

if __name__ == "__main__":

    print("=" * 70)
    print("Peer Comparison Excel Report")
    print("=" * 70)

    conn = sqlite3.connect(DB_PATH)

    # ------------------------------------------------------
    # Load peer-group companies
    # ------------------------------------------------------

    peer_groups = pd.read_sql(
        """
        SELECT
            peer_group_name,
            company_id,
            is_benchmark
        FROM peer_groups
        """,
        conn,
    )

    # ------------------------------------------------------
    # Company names
    # companies.id is the company identifier
    # ------------------------------------------------------

    companies = pd.read_sql(
        """
        SELECT
            id AS company_id,
            company_name
        FROM companies
        """,
        conn,
    )

    # ------------------------------------------------------
    # Financial ratios
    # ------------------------------------------------------

    ratios = pd.read_sql(
        """
        SELECT
            company_id,
            year,
            return_on_equity_pct,
            return_on_capital_employed_pct,
            net_profit_margin_pct,
            operating_profit_margin_pct,
            debt_to_equity,
            interest_coverage,
            asset_turnover,
            free_cash_flow_cr,
            cash_from_operations_cr,
            revenue_cagr_5yr,
            pat_cagr_5yr,
            eps_cagr_5yr,
            composite_quality_score
        FROM financial_ratios
        """,
        conn,
    )

    # ------------------------------------------------------
    # Market data
    # ------------------------------------------------------

    market = pd.read_sql(
        """
        SELECT
            company_id,
            year,
            market_cap_crore,
            pe_ratio,
            pb_ratio,
            dividend_yield_pct
        FROM market_cap
        """,
        conn,
    )

    # ------------------------------------------------------
    # Profit & Loss
    # ------------------------------------------------------

    pnl = pd.read_sql(
        """
        SELECT
            company_id,
            year,
            sales,
            net_profit,
            dividend_payout
        FROM profitandloss
        """,
        conn,
    )

    # ------------------------------------------------------
    # Day 18 percentile table
    # ------------------------------------------------------

    peer_pct = pd.read_sql(
        """
        SELECT
            company_id,
            peer_group_name,
            metric,
            value,
            percentile_rank,
            year
        FROM peer_percentiles
        """,
        conn,
    )

    conn.close()

    # ======================================================
    # Prepare financial data
    # ======================================================

    # Convert year to string
    ratios["year"] = ratios["year"].astype(str)
    market["year"] = market["year"].astype(str)
    pnl["year"] = pnl["year"].astype(str)

    # ------------------------------------------------------
    # Remove exact duplicate rows
    # ------------------------------------------------------

    ratios = ratios.drop_duplicates()

    market = market.drop_duplicates()

    pnl = pnl.drop_duplicates()

    # ------------------------------------------------------
    # Select latest available financial year per company
    # ------------------------------------------------------

    ratios["_year_num"] = pd.to_numeric(
        ratios["year"].str[-4:],
        errors="coerce",
    )

    ratios = (
        ratios.sort_values(
            ["company_id", "_year_num"],
            ascending=[True, False],
        )
        .drop_duplicates(
            "company_id",
            keep="first",
        )
        .drop(columns="_year_num")
    )

    # ------------------------------------------------------
    # Merge market data
    # ------------------------------------------------------

    ratios["_year_num"] = pd.to_numeric(
        ratios["year"].str[-4:],
        errors="coerce",
    )

    market["_year_num"] = pd.to_numeric(
        market["year"].str[-4:],
        errors="coerce",
    )

    market = market.sort_values(
        ["company_id", "_year_num"],
        ascending=[True, False],
    ).drop_duplicates(
        "company_id",
        keep="first",
    )

    ratios = ratios.merge(
        market[
            [
                "company_id",
                "market_cap_crore",
                "pe_ratio",
                "pb_ratio",
                "dividend_yield_pct",
            ]
        ],
        on="company_id",
        how="left",
    )

    # ------------------------------------------------------
    # Merge P&L
    # ------------------------------------------------------

    pnl["_year_num"] = pd.to_numeric(
        pnl["year"].str[-4:],
        errors="coerce",
    )

    pnl = pnl.sort_values(
        ["company_id", "_year_num"],
        ascending=[True, False],
    ).drop_duplicates(
        "company_id",
        keep="first",
    )

    ratios = ratios.merge(
        pnl[
            [
                "company_id",
                "sales",
                "net_profit",
                "dividend_payout",
            ]
        ],
        on="company_id",
        how="left",
    )

    # ------------------------------------------------------
    # Add company names
    # ------------------------------------------------------

    ratios = ratios.merge(
        companies,
        on="company_id",
        how="left",
    )

    # ------------------------------------------------------
    # Add peer-group information
    # ------------------------------------------------------

    ratios = ratios.merge(
        peer_groups,
        on="company_id",
        how="inner",
    )

    # ======================================================
    # Calculate percentile ranks for the 20 metrics
    # within each peer group
    # ======================================================

    for metric in METRICS:

        ratios[metric] = pd.to_numeric(
            ratios[metric],
            errors="coerce",
        )

        ratios[f"{metric}_percentile"] = ratios.groupby(
            "peer_group_name", group_keys=False
        )[metric].transform(
            lambda x, metric=metric: calculate_percentile(
                x,
                inverse=metric in LOWER_IS_BETTER,
            )
        )

    # ======================================================
    # Create Excel workbook
    # ======================================================

    writer = pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    )

    groups = sorted(ratios["peer_group_name"].dropna().unique())

    print(f"\nPeer groups found: {len(groups)}")

    # ======================================================
    # Create one sheet per peer group
    # ======================================================

    for group in groups:

        sheet = ratios[ratios["peer_group_name"] == group].copy()

        # --------------------------------------------------
        # Sort benchmark first, then company
        # --------------------------------------------------

        sheet["_benchmark_sort"] = ~sheet["is_benchmark"].astype(bool)

        sheet = sheet.sort_values(
            [
                "_benchmark_sort",
                "company_id",
            ]
        ).drop(columns="_benchmark_sort")

        # --------------------------------------------------
        # Select required columns
        # --------------------------------------------------

        output_columns = [
            "company_id",
            "company_name",
        ]

        # 20 metric columns
        output_columns += METRICS

        # Percentile rank for every metric
        percentile_columns = [f"{metric}_percentile" for metric in METRICS]

        output_columns += percentile_columns

        # Benchmark flag is needed internally for formatting
        output_columns += ["is_benchmark"]

        output = sheet[[col for col in output_columns if col in sheet.columns]].copy()

        # --------------------------------------------------
        # Write sheet
        # --------------------------------------------------

        output.to_excel(
            writer,
            sheet_name=str(group)[:31],
            index=False,
        )

        print(f"Created sheet: {group} " f"({len(output)} companies)")

    writer.close()

    # ======================================================
    # Format workbook
    # ======================================================

    wb = load_workbook(OUTPUT_FILE)

    for ws in wb.worksheets:

        headers = [cell.value for cell in ws[1]]

        # --------------------------------------------------
        # Freeze header
        # --------------------------------------------------

        ws.freeze_panes = "A2"

        # --------------------------------------------------
        # Header formatting
        # --------------------------------------------------

        for cell in ws[1]:

            cell.font = Font(bold=True)

            cell.alignment = Alignment(horizontal="center")

        # --------------------------------------------------
        # Find benchmark column
        # --------------------------------------------------

        if "is_benchmark" in headers:

            benchmark_col = headers.index("is_benchmark") + 1

        else:

            benchmark_col = None

        # --------------------------------------------------
        # Colour percentile cells
        # --------------------------------------------------

        for col in range(
            1,
            ws.max_column + 1,
        ):

            header = ws.cell(
                1,
                col,
            ).value

            if header is None or not str(header).endswith("_percentile"):
                continue

            for row in range(
                2,
                ws.max_row + 1,
            ):

                value = ws.cell(
                    row,
                    col,
                ).value

                if value is None:
                    continue

                try:
                    value = float(value)
                except (TypeError, ValueError):
                    continue

                # Green >= 75%
                if value >= 0.75:

                    ws.cell(
                        row,
                        col,
                    ).fill = GREEN

                # Red <= 25%
                elif value <= 0.25:

                    ws.cell(
                        row,
                        col,
                    ).fill = RED

                # Yellow 25% - 75%
                else:

                    ws.cell(
                        row,
                        col,
                    ).fill = YELLOW

        # --------------------------------------------------
        # Highlight benchmark row
        # --------------------------------------------------

        if benchmark_col is not None:

            for row in range(
                2,
                ws.max_row + 1,
            ):

                benchmark = ws.cell(
                    row,
                    benchmark_col,
                ).value

                if benchmark in (
                    1,
                    True,
                    "1",
                    "True",
                ):

                    for col in range(
                        1,
                        ws.max_column + 1,
                    ):

                        ws.cell(
                            row,
                            col,
                        ).fill = GOLD

                        ws.cell(
                            row,
                            col,
                        ).font = Font(bold=True)

        # --------------------------------------------------
        # Add peer median row
        # --------------------------------------------------

        median_row = ws.max_row + 2

        ws.cell(
            median_row,
            1,
        ).value = "Peer Median"

        ws.cell(
            median_row,
            1,
        ).font = Font(bold=True)

        # Find metric columns
        for metric in METRICS:

            if metric not in headers:
                continue

            col = headers.index(metric) + 1

            values = []

            for row in range(
                2,
                ws.max_row + 1,
            ):

                value = ws.cell(
                    row,
                    col,
                ).value

                if value is not None:

                    try:
                        values.append(float(value))
                    except (
                        TypeError,
                        ValueError,
                    ):
                        pass

            if values:

                ws.cell(
                    median_row,
                    col,
                ).value = round(
                    pd.Series(values).median(),
                    2,
                )

        # --------------------------------------------------
        # Style median row
        # --------------------------------------------------

        for col in range(
            1,
            ws.max_column + 1,
        ):

            ws.cell(
                median_row,
                col,
            ).font = Font(bold=True)

        # --------------------------------------------------
        # Column widths
        # --------------------------------------------------

        for column_cells in ws.columns:

            max_length = 0

            column_letter = column_cells[0].column_letter

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            ws.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                28,
            )

    # ------------------------------------------------------
    # Save final workbook
    # ------------------------------------------------------

    wb.save(OUTPUT_FILE)

    print("\n" + "=" * 70)
    print("Peer comparison report created successfully.")
    print("=" * 70)
    print(f"Output: {OUTPUT_FILE}")
    print(f"Sheets created: {len(wb.sheetnames)}")
    print(
        "Sheets:",
        ", ".join(wb.sheetnames),
    )
    print("=" * 70)
